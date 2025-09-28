from textual.app import App, ComposeResult
from textual.widgets import Static, Button, Input, Label
from textual.containers import Vertical, Horizontal, Container
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws["A2"] = "Nombre:"
ws["A3"] = "Edad:"
ws["A4"] = "Email:"


class MiApp(App):
    CSS_PATH = "./style.tcss"

    def compose(self) -> ComposeResult:
        with Container():
            with Vertical():
                yield Static("Formulario de prueba")

                yield Label("Nombre:")
                yield Input(placeholder="Ingresá tu nombre", id="nombre")

                yield Label("Edad:")
                yield Input(placeholder="Ingresá tu edad", id="edad")

                yield Label("Email:")
                yield Input(placeholder="Ingresá tu email", id="email")

                yield Button("Guardar datos", id="guardar", classes="boton-principal")

                yield Static("", id="salida")

    def on_button_pressed(self, event: Button.Pressed) -> None:
        global ws, wb
        if event.button.id == "guardar":
            nombre = self.query_one("#nombre", Input).value
            edad = self.query_one("#edad", Input).value
            email = self.query_one("#email", Input).value

            ws["B2"] = nombre
            ws["B3"] = edad
            ws["B4"] = email

            wb.save(nombre + ".xlsx")

            salida = self.query_one("#salida", Static)
            salida.update(f"El archivo {nombre}.xlsx ha sido creado correctamente.")


if __name__ == "__main__":
    app = MiApp()
    app.run()
