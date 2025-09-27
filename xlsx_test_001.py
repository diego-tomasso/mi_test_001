from openpyxl import Workbook

# Crear un nuevo archivo Excel en memoria
wb = Workbook()
ws = wb.active  # Selecciona la hoja activa (por defecto se llama "Sheet")
# Agregar una nueva hoja con nombre personalizado
w2 = wb.create_sheet(title="ShLeti")

# Escribir el valor 1 en la celda A1
ws["A1"] = 1
ws["A2"] = "Gastos en camino"

# Escribir en la sheet llamada "ShLeti" las siguientes celdas
w2["A1"] = 31
w2["B2"] = "Texto!"
w2["A2"] = 19
w2["A3"] = "=A1+A2"

# Guardard el archivo
wb.save("ejemplo.xlsx")
