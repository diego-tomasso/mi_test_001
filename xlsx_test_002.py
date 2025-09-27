from openpyxl import load_workbook

# Cargar el archivo existente
wb = load_workbook("ejemplo.xlsx")

# Acceder a la hoja activa
ws = wb.active

# Acceder a una hoja por nombre
w2 = wb["ShLeti"]

# Leer el valor de la celda A1
valor_w2_a1 = w2["A1"].value
print(f"Valor original en ShLeti A1: {valor_w2_a1}.")
valor_w2_a1 = int(valor_w2_a1)
valor_w2_a1 += 1
# Modificar el valor de la celda SheLeti[A1]
w2["A1"] = valor_w2_a1

# Leer e imprimir un texto desde el file
valor_ws_a2 = ws["A2"].value
print(f"El texto en Sheet[A2] es: {valor_ws_a2}.")

# Guarda los cambios (puede sobrescribir el archivo original)
wb.save("ejemplo.xlsx")
